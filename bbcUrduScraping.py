from bs4 import BeautifulSoup
import requests
import pandas as pd
from openpyxl import load_workbook
from collections import Counter
import matplotlib.pyplot as plt
from tkinter import *

def main(UserInput):
    def hl_with_story(url, count, category):
        category_link = requests.get(url).text
        soup2 = BeautifulSoup(category_link, 'lxml')
        category_headlines = soup2.find_all('h2', class_='bbc-4z1it4 ewr0zp70')

        for ch in category_headlines:
            story = ''
            count += 1
            if  count <= 100:
                h_link = ch.a['href']
                headline_link = requests.get(h_link).text
                soup3 = BeautifulSoup(headline_link, 'lxml')
                mazeedPrhye = soup3.find_all('div', class_='bbc-4wucq3 essoxwk0')
                size = len(mazeedPrhye)
                i = 0
                while i < size:
                    story = story + mazeedPrhye[i].text
                    i+=1
                df = pd.DataFrame({'Story':[story],
                                   'Headline':[ch.text],
                                   'Category':[category]})
                df.to_csv('BBC.csv', mode='a', index=False, header=False)
        return count

    html_file_bbcUrdu = requests.get(UserInput).text
    soup = BeautifulSoup(html_file_bbcUrdu, 'lxml')
    categories = soup.find_all('li', class_='bbc-22iqn0 e1ibkbh72')
    for category in categories:
        count = 0
        if category.text != str('موجودہ صفحہ, صفحۂ اول') and category.text != str('ویڈیو'):
            no = 1
            link = category.a['href']
            url = 'https://www.bbc.com' + link + '?page=' + str(no)
            # function call 1
            count = hl_with_story(url, count, category.text)
            while count%24==0 and count <= 100:
                no += 1
                url = 'https://www.bbc.com' + link + '?page=' + str(no)
                # function call 2
                count = hl_with_story(url, count, category.text)

def wordsList():
    dataSet = 'BBC.xlsx'
    wb = load_workbook(dataSet)
    ws = wb['Sheet 2']
    all_cols = list(ws.columns)
    store = ''
    for cell in all_cols[0]:
        store = store + ' ' + cell.value
    newString = ""
    for i in store:
        if i.isalnum() or i.isspace():
            newString += i
    words = newString.split()
    return words

def MaxStory():
    dataSet = 'BBC.xlsx'
    wb = load_workbook(dataSet)
    ws = wb['Sheet 2']
    all_cols = list(ws.columns)
    max = 0
    maxStory = ''
    for story in all_cols[0]:
        if len(story.value) != 0 and story.value != 'Story':
            if len(story.value) > max:
                max = len(story.value)
                maxStory = story.value
    str1 = 'Maximum Length of Story: ' + str(max) + '\n' + str(maxStory)
    newWindow = Tk()
    newWindow.title('Maximum Story')
    newWindow.geometry('683x384+550+180')
    newWindow.resizable(False, False)

    frame = Frame(newWindow, width=683, height=384, bg='#181B22')
    frame.place(relx=0.5, rely=0.5, anchor='center')

    v = Scrollbar(newWindow, orient='vertical')
    v.pack(side=RIGHT, fill='y')

    h = Scrollbar(newWindow, orient='horizontal')
    h.pack(side=BOTTOM, fill='x')
    text = Text(frame, height=16, width=65, font=('Times New Roman', 15), yscrollcommand=v.set)
    v.config(command=text.yview)
    h.config(command=text.xview)
    text.place(relx=0.5, rely=0.5, anchor='center')
    text.insert(END, str1)

    newWindow.mainloop()

def MinStory():
    dataSet = 'BBC.xlsx'
    wb = load_workbook(dataSet)
    ws = wb['Sheet 2']
    all_cols = list(ws.columns)
    min = 2**31-1
    minStory = ''
    for story in all_cols[0]:
        if len(story.value) != 0 and story.value != 'Story':
            if len(story.value) < min:
                min = len(story.value)
                minStory = story.value
    str1 = 'Minimum Length of Story: ' + str(min) + '\n' + str(minStory)
    newWindow = Tk()
    newWindow.title('Minimum Story')
    newWindow.geometry('683x384+550+180')
    newWindow.resizable(False, False)

    frame = Frame(newWindow, width=683, height=384, bg='#181B22')
    frame.place(relx=0.5, rely=0.5, anchor='center')

    v = Scrollbar(newWindow, orient='vertical')
    v.pack(side=RIGHT, fill='y')

    h = Scrollbar(newWindow, orient='horizontal')
    h.pack(side=BOTTOM, fill='x')
    text = Text(frame, height=16, width=65, font=('Times New Roman', 15), yscrollcommand=v.set)
    v.config(command=text.yview)
    h.config(command=text.xview)
    text.place(relx=0.5, rely=0.5, anchor='center')
    text.insert(END, str1)

    newWindow.mainloop()

def wordsFrequency():
    all_words = wordsList()
    words_count = Counter(all_words)
    mostOccur = words_count.most_common(10)

    newWindow = Tk()
    newWindow.title('Top 10 words in terms of FREQUENCY')
    newWindow.geometry('683x384+550+180')
    newWindow.resizable(False, False)

    frame = Frame(newWindow, width=683, height=384, bg='#181B22')
    frame.place(relx=0.5, rely=0.5, anchor='center')

    str1 = 'Top 10 words in terms of FREQUENCY: \n\n'
    for i in mostOccur:
        str1 = str1 + str(i) + '\n'

    v = Scrollbar(newWindow, orient='vertical')
    v.pack(side=RIGHT, fill='y')

    h = Scrollbar(newWindow, orient='horizontal')
    h.pack(side=BOTTOM, fill='x')

    text = Text(frame, height=16, width=65, font=('Times New Roman', 15), yscrollcommand=v.set)
    v.config(command=text.yview)
    h.config(command=text.xview)
    text.place(relx=0.5, rely=0.5, anchor='center')
    text.insert(END, str1)

    newWindow.mainloop()

def countUniqueWords():
    words = wordsList()
    List = []

    newWindow = Tk()
    newWindow.title('Unique Words')
    newWindow.geometry('683x384+550+180')
    newWindow.resizable(False, False)

    frame = Frame(newWindow, width=683, height=384, bg='#181B22')
    frame.place(relx=0.5, rely=0.5, anchor='center')

    for i in words:
        if i != 'Story':
            if i not in List:
                List.append(i)
            else:
                continue
    Single_string = 'Total Words: ' + str(len(words)) + '\n' + 'Unique Words: ' + str(len(List)) + '\n'
    i = 0
    while i < len(List):
        Single_string = Single_string + List[i] + '\n'
        i+=1

    v = Scrollbar(newWindow, orient='vertical')
    v.pack(side=RIGHT, fill='y')

    h = Scrollbar(newWindow, orient='horizontal')
    h.pack(side=BOTTOM, fill='x')

    text = Text(frame, height=16, width=65, font=('Times New Roman', 15))
    text.place(relx=0.5, rely=0.5, anchor='center')
    v.config(command=text.yview)
    h.config(command=text.xview)
    text.insert(END, Single_string)

    newWindow.mainloop()

def countStories():
    List = [0,0,0,0,0,0]
    dataSet = 'BBC.xlsx'
    wb = load_workbook(dataSet)
    ws = wb['Sheet 2']
    all_cols = list(ws.columns)
    for i in all_cols[2]:
        if i.value == 'پاکستان':
            List[0] += 1
        elif i.value == 'آس پاس':
            List[1] += 1
        elif i.value == 'ورلڈ':
            List[2] += 1
        elif i.value == 'کھیل':
            List[3] += 1
        elif i.value == 'فن فنکار':
            List[4] += 1
        elif i.value == 'سائنس':
            List[5] += 1
    return List

def printCountStories():
    List = countStories()
    str1 = 'Stories in each category: \n' + \
           str(List[0]) + ' Stories retrieved from PAKISTAN Category \n' + \
           str(List[1]) + ' Stories retrieved from AASPAS Category \n' + \
           str(List[2]) + ' Stories retrieved from WORLD Category \n' + \
           str(List[3]) + ' Stories retrieved from KHEL Category \n' + \
           str(List[4]) + ' Stories retrieved from FANKAR Category \n' + \
           str(List[5]) + ' Stories retrieved from SCIENCE Category'

    newWindow = Tk()
    newWindow.title('Stories In Each Category')
    newWindow.geometry('683x384+550+180')
    newWindow.resizable(False, False)

    frame = Frame(newWindow, width=683, height=384, bg='#181B22')
    frame.place(relx=0.5, rely=0.5, anchor='center')

    v = Scrollbar(newWindow, orient='vertical')
    v.pack(side=RIGHT, fill='y')

    h = Scrollbar(newWindow, orient='horizontal')
    h.pack(side=BOTTOM, fill='x')

    text = Text(frame, height=16, width=65, font=('Times New Roman', 15))
    text.place(relx=0.5, rely=0.5, anchor='center')
    v.config(command=text.yview)
    h.config(command=text.xview)
    text.insert(END, str1)
    newWindow.mainloop()

def barGraph():
    List = countStories()
    catList = ['Pakistan', 'AasPas', 'World', 'Khel', 'Fankar', 'Science']
    plt.bar(catList, List)
    plt.xlabel('Categories')
    plt.ylabel('No. of Stories')
    plt.title('Graph showing no. of records for each categroy')
    plt.show()

def GUI():
    mainWindow = Tk()
    mainWindow.title('Web Scraping')
    mainWindow.geometry('800x768+0+0')
    mainWindow.resizable(False, False)

    frame = Frame(mainWindow, width=800, height=768, bg='#181B22')
    frame.place(relx=0.5, rely=0.5, anchor='center')

    label1 = Label(frame, text='Enter URL', font=('TImes New Roman', 20), bg='#181B22', fg='white')
    label1.place(relx=0.18, rely=0.1, relwidth=0.15, relheight=0.06, anchor='center')

    url = Entry(frame, font=('Times New Roman', 20), bd=7)
    url.place(relx=0.48, rely=0.1, relwidth=0.38, relheight=0.06, anchor='center')

    scrapButton = Button(frame, text='Scrap', font=('Times New Roman', 20), command=lambda: main(url.get()), bd=7)
    scrapButton.place(relx=0.70, rely=0.07, relwidth=0.22, relheight=0.06)

    Button1 = Button(frame, text='Unique Words', font=('Times New Roman', 20), command=countUniqueWords, bd=7)
    Button1.place(relx=0.5, rely=0.23, relwidth=0.23, relheight=0.06, anchor='center')

    Button2 = Button(frame, text='Count Stories', font=('Times New Roman', 20), command=printCountStories, bd=7)
    Button2.place(relx=0.5, rely=0.36, relwidth=0.23, relheight=0.06, anchor='center')

    Button3 = Button(frame, text='Max Story', font=('Times New Roman', 20), command=MaxStory, bd=7)
    Button3.place(relx=0.5, rely=0.49, relwidth=0.23, relheight=0.06, anchor='center')

    Button4 = Button(frame, text='Min Story', font=('Times New ROman', 20), command=MinStory, bd=7)
    Button4.place(relx=0.5, rely=0.62, relwidth=0.23, relheight=0.06, anchor='center')

    Button5 = Button(frame, text='Top 10 words', font=('Times New Roman', 20), command=wordsFrequency, bd=7)
    Button5.place(relx=0.5, rely=0.75, relwidth=0.23, relheight=0.06, anchor='center')

    Button6 = Button(frame, text='Bar Graph', font=('Times New Roman', 20), command=barGraph, bd=10)
    Button6.place(relx=0.5, rely=0.88, relwidth=0.23, relheight=0.06, anchor='center')

    mainWindow.mainloop()



GUI()